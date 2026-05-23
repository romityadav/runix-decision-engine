"""Load and validate the Excel scenario workbook into the typed domain model.

This loader does three things that the two earlier prototypes did not:

1. It reads the *named constants* as the single source of truth and derives each
   order's SLA window and breach penalty from its tier — because the workbook's
   per-row formula columns contain a cell-reference bug (the "Breach Penalty"
   column points at the SLA-hours cells, and the "SLA Window" column ignores the
   tier entirely). See ``docs/DATA_NOTES.md`` for the full audit.

2. It never trusts formula-derived cells for values: this workbook was saved
   without a recalculation pass, so every formula cell reads back as ``None``.
   All driver economics therefore come from the constants block, and aggregate
   totals are computed here, from the rows.

3. It records every data-quality observation on the returned ``Scenario`` and
   logs it at WARNING, so the findings travel with the data instead of being
   silently swallowed.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .logging_setup import get_logger
from .models import (
    Driver,
    Order,
    Scenario,
    ScenarioConstants,
    ScenarioContext,
    ServiceTier,
)

log = get_logger("runix.data_loader")

# Sheet names as they appear in the workbook.
_SCENARIO_SHEET = "Scenario Summary"
_ORDERS_SHEET = "Orders Data"
_DRIVERS_SHEET = "Integrated Drivers"


class DataLoaderError(RuntimeError):
    """Raised when the workbook cannot be parsed into a valid scenario."""


def load_scenario(path: str | Path) -> Scenario:
    """Parse the workbook at *path* into a fully typed :class:`Scenario`.

    Raises:
        FileNotFoundError: if the path does not exist.
        DataLoaderError: if a required sheet, constant, or column is missing.
    """
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Dataset not found: {workbook_path}")

    try:
        wb = load_workbook(workbook_path, data_only=True)
    except Exception as exc:  # openpyxl raises a grab-bag of exceptions
        raise DataLoaderError(f"Could not open workbook '{workbook_path}': {exc}") from exc

    for sheet in (_SCENARIO_SHEET, _ORDERS_SHEET, _DRIVERS_SHEET):
        if sheet not in wb.sheetnames:
            raise DataLoaderError(
                f"Required sheet '{sheet}' missing. Found: {wb.sheetnames}"
            )

    notes: list[str] = []
    constants = _load_constants(wb[_SCENARIO_SHEET])
    context = _load_context(wb[_SCENARIO_SHEET])
    orders = _load_orders(wb[_ORDERS_SHEET], constants, notes)
    drivers = _load_drivers(wb[_DRIVERS_SHEET], notes)

    _sanity_check(constants, context, orders, drivers, notes)

    for note in notes:
        log.warning("data-quality: %s", note)

    return Scenario(
        constants=constants,
        context=context,
        orders=orders,
        drivers=drivers,
        data_quality_notes=notes,
    )


# ---------------------------------------------------------------------------
# Scenario Summary sheet
# ---------------------------------------------------------------------------


def _scan_key_value(sheet: Any, key_col: int, val_col: int) -> dict[str, Any]:
    """Collect ``{label: value}`` pairs from a two-column key/value block."""
    pairs: dict[str, Any] = {}
    for row in sheet.iter_rows(values_only=True):
        if len(row) < max(key_col, val_col) + 1:
            continue
        key, val = row[key_col], row[val_col]
        if key is not None and val is not None:
            pairs[str(key).strip()] = val
    return pairs


def _require(pairs: dict[str, Any], key: str, sheet: str) -> Any:
    if key not in pairs:
        raise DataLoaderError(f"Required parameter '{key}' missing from sheet '{sheet}'.")
    return pairs[key]


def _load_constants(sheet: Any) -> ScenarioConstants:
    # Engine constants live in columns B (label, index 1) / C (value, index 2).
    c = _scan_key_value(sheet, key_col=1, val_col=2)
    return ScenarioConstants(
        shift_hours=int(_require(c, "SHIFT_HOURS", _SCENARIO_SHEET)),
        deliveries_per_driver_per_shift=float(
            _require(c, "DELIVERIES_PER_DRIVER_PER_SHIFT", _SCENARIO_SHEET)
        ),
        express_sla_hours=float(_require(c, "EXPRESS_SLA_HOURS", _SCENARIO_SHEET)),
        standard_sla_hours=float(_require(c, "STANDARD_SLA_HOURS", _SCENARIO_SHEET)),
        driver_shift_cost=float(_require(c, "DRIVER_SHIFT_COST", _SCENARIO_SHEET)),
        express_penalty=float(_require(c, "EXPRESS_PENALTY", _SCENARIO_SHEET)),
        standard_penalty=float(_require(c, "STANDARD_PENALTY", _SCENARIO_SHEET)),
        weather_multiplier=float(
            _require(c, "WEATHER_MULTIPLIER (heavy_snow)", _SCENARIO_SHEET)
        ),
        traffic_risk_score=float(_require(c, "TRAFFIC_RISK_SCORE (major)", _SCENARIO_SHEET)),
        weather_risk_score=float(_require(c, "WEATHER_RISK_SCORE (heavy_snow)", _SCENARIO_SHEET)),
        risk_weights=_parse_weights(
            _require(c, "RISK_WEIGHTS (weather/traffic/load)", _SCENARIO_SHEET)
        ),
    )


def _load_context(sheet: Any) -> ScenarioContext:
    # Active scenario context lives in columns F (label, index 5) / G (value, index 6).
    c = _scan_key_value(sheet, key_col=5, val_col=6)
    return ScenarioContext(
        weather_condition=str(_require(c, "Current Weather Condition", _SCENARIO_SHEET)).strip(),
        traffic_severity=str(
            _require(c, "Active Traffic/Event Severity", _SCENARIO_SHEET)
        ).strip(),
        target_order_volume=int(_require(c, "Target Order Volume", _SCENARIO_SHEET)),
        base_shift_drivers_available=int(
            _require(c, "Base Shift Drivers Available", _SCENARIO_SHEET)
        ),
        express_tier_share=_parse_percent(
            _require(c, "Express Order Tier Mix Share", _SCENARIO_SHEET)
        ),
        primary_affected_zone=str(_require(c, "Primary Affected Zone", _SCENARIO_SHEET))
        .strip()
        .lower(),
        affected_zone_share=_parse_percent(_require(c, "Affected Zone Share", _SCENARIO_SHEET)),
    )


# ---------------------------------------------------------------------------
# Orders + Drivers sheets
# ---------------------------------------------------------------------------


def _load_orders(sheet: Any, constants: ScenarioConstants, notes: list[str]) -> list[Order]:
    """Read orders from the hand-entered id/tier/zone columns.

    SLA hours and penalties are derived from the named constants by tier — we do
    NOT read the per-row formula columns E/F, which are buggy. We record that
    decision once, here.
    """
    notes.append(
        "Orders sheet columns 'SLA Delivery Window' and 'Breach Penalty Risk' contain a "
        "cell-reference bug (penalty column points at SLA-hours cells; SLA column ignores "
        "tier). Deriving SLA hours and penalties from the named constants instead."
    )

    orders: list[Order] = []
    for row in sheet.iter_rows(values_only=True):
        if len(row) < 4:
            continue
        oid = row[1]  # column B
        if not (isinstance(oid, str) and oid.startswith("ORD-")):
            continue  # skips title rows, header row, and the "Total ..." footer
        tier_raw = str(row[2]).strip().lower()  # column C
        zone = str(row[3]).strip().lower()  # column D
        try:
            tier = ServiceTier(tier_raw)
        except ValueError:
            notes.append(f"Order {oid} has unknown tier '{tier_raw}'; treating as standard.")
            tier = ServiceTier.STANDARD
        is_express = tier is ServiceTier.EXPRESS
        sla = constants.express_sla_hours if is_express else constants.standard_sla_hours
        penalty = constants.express_penalty if is_express else constants.standard_penalty
        orders.append(
            Order(order_id=oid.strip(), tier=tier, zone=zone, sla_hours=sla, breach_penalty=penalty)
        )
    if not orders:
        raise DataLoaderError(f"No orders found in sheet '{_ORDERS_SHEET}'.")
    return orders


def _load_drivers(sheet: Any, notes: list[str]) -> list[Driver]:
    """Read drivers from the hand-entered id/status columns only.

    Numeric driver columns (shift cost, degraded capacity) are workbook formulas
    saved without cached values, so they read as ``None``; driver economics come
    from the constants block instead.
    """
    drivers: list[Driver] = []
    for row in sheet.iter_rows(values_only=True):
        if len(row) < 3:
            continue
        did = row[1]  # column B
        if not (isinstance(did, str) and did.startswith("DRV-")):
            continue
        drivers.append(Driver(driver_id=did.strip(), status=str(row[2]).strip()))
    if not drivers:
        raise DataLoaderError(f"No drivers found in sheet '{_DRIVERS_SHEET}'.")
    return drivers


def _sanity_check(
    constants: ScenarioConstants,
    context: ScenarioContext,
    orders: list[Order],
    drivers: list[Driver],
    notes: list[str],
) -> None:
    """Cross-check the data against itself and against reality, recording drift.

    This is the "does this make sense?" pass. None of these are fatal — they are
    exactly the kind of quiet inconsistency that, left unchecked, lets a model
    silently learn garbage.
    """
    n = len(orders)
    if n != context.target_order_volume:
        notes.append(
            f"Order rows ({n}) != Target Order Volume ({context.target_order_volume})."
        )
    active = sum(1 for d in drivers if d.is_active)
    if active != context.base_shift_drivers_available:
        notes.append(
            f"Active drivers ({active}) != Base Shift Drivers Available "
            f"({context.base_shift_drivers_available})."
        )
    express = sum(1 for o in orders if o.is_express)
    observed_share = express / n if n else 0.0
    if abs(observed_share - context.express_tier_share) > 0.01:
        notes.append(
            f"Express share in rows ({observed_share:.0%}) differs from stated "
            f"Express Order Tier Mix Share ({context.express_tier_share:.0%})."
        )
    affected = sum(1 for o in orders if o.zone == context.primary_affected_zone)
    observed_zone_share = affected / n if n else 0.0
    if abs(observed_zone_share - context.affected_zone_share) > 0.01:
        notes.append(
            f"Affected-zone ('{context.primary_affected_zone}') share in rows "
            f"({observed_zone_share:.0%}) differs from stated Affected Zone Share "
            f"({context.affected_zone_share:.0%}); using the row-level counts for capacity."
        )
    if not 0.0 < constants.weather_multiplier <= 1.0:
        notes.append(
            f"WEATHER_MULTIPLIER {constants.weather_multiplier} outside (0, 1]; "
            "expected a capacity-reducing multiplier."
        )


# ---------------------------------------------------------------------------
# Small parsers
# ---------------------------------------------------------------------------


def _parse_percent(value: Any) -> float:
    """Parse '40.0%' or 40 or 0.4 into a 0-1 fraction."""
    if isinstance(value, str) and value.strip().endswith("%"):
        return float(value.strip().rstrip("%")) / 100.0
    num = float(value)
    return num / 100.0 if num > 1.0 else num


def _parse_weights(value: Any) -> dict[str, float]:
    """Parse '40% / 30% / 30%' into ``{weather, traffic, load}`` fractions."""
    default = {"weather": 0.4, "traffic": 0.3, "load": 0.3}
    if not isinstance(value, str):
        return default
    parts = [p.strip() for p in value.split("/")]
    if len(parts) != 3:
        return default
    return {
        "weather": _parse_percent(parts[0]),
        "traffic": _parse_percent(parts[1]),
        "load": _parse_percent(parts[2]),
    }
