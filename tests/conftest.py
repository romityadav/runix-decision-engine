"""Shared fixtures: the real dataset, a loaded scenario, and a synthetic-workbook
builder for exercising the loader's validation paths.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from runix.config import EngineConfig
from runix.data_loader import load_scenario
from runix.models import Scenario

_DATA_DIR = Path(__file__).resolve().parent.parent / "data"
DATASET = _DATA_DIR / "Runix_Logistics_Engine_Scenario_Dataset.xlsx"


@pytest.fixture(scope="session")
def dataset_path() -> Path:
    assert DATASET.exists(), f"bundled dataset missing at {DATASET}"
    return DATASET


@pytest.fixture
def scenario(dataset_path: Path) -> Scenario:
    return load_scenario(dataset_path)


@pytest.fixture
def config() -> EngineConfig:
    return EngineConfig()


# --- synthetic workbook builder ------------------------------------------------

_REQUIRED_CONSTANTS = {
    "SHIFT_HOURS": 8,
    "DELIVERIES_PER_DRIVER_PER_SHIFT": 8,
    "EXPRESS_SLA_HOURS": 2,
    "STANDARD_SLA_HOURS": 8,
    "DRIVER_SHIFT_COST": 120,
    "EXPRESS_PENALTY": 25,
    "STANDARD_PENALTY": 8,
    "WEATHER_MULTIPLIER (heavy_snow)": 0.6,
    "TRAFFIC_RISK_SCORE (major)": 75,
    "WEATHER_RISK_SCORE (heavy_snow)": 90,
    "RISK_WEIGHTS (weather/traffic/load)": "40% / 30% / 30%",
}

_CONTEXT = {
    "Current Weather Condition": "heavy_snow",
    "Active Traffic/Event Severity": "major",
    "Target Order Volume": 6,
    "Base Shift Drivers Available": 2,
    "Express Order Tier Mix Share": "50.0%",
    "Primary Affected Zone": "central",
    "Affected Zone Share": "50.0%",
}


@pytest.fixture
def make_workbook():
    """Return the synthetic-workbook builder (factory fixture)."""
    return _build_workbook


def _build_workbook(
    path: Path,
    *,
    omit_sheet: str | None = None,
    omit_param: str | None = None,
    n_orders: int = 6,
    n_drivers: int = 2,
) -> Path:
    """Write a small but structurally valid workbook, optionally breaking it."""
    wb = Workbook()
    wb.remove(wb.active)

    if omit_sheet != "Scenario Summary":
        ss = wb.create_sheet("Scenario Summary")
        ss["B2"] = "Scenario Data"
        ss["B5"], ss["C5"] = "Parameter", "Value"
        ss["F5"], ss["G5"] = "Variable", "Config Value"
        r = 6
        for k, v in _REQUIRED_CONSTANTS.items():
            if k == omit_param:
                continue
            ss.cell(r, 2, k)
            ss.cell(r, 3, v)
            r += 1
        r = 6
        for k, v in _CONTEXT.items():
            if k == omit_param:
                continue
            ss.cell(r, 6, k)
            ss.cell(r, 7, v)
            r += 1

    if omit_sheet != "Orders Data":
        od = wb.create_sheet("Orders Data")
        od["B4"], od["C4"], od["D4"] = "Order ID", "Service Tier", "Destination Zone"
        for i in range(n_orders):
            tier = "express" if i % 2 == 0 else "standard"
            zone = "central" if i % 2 == 0 else "north"
            od.cell(5 + i, 2, f"ORD-{i:03d}")
            od.cell(5 + i, 3, tier)
            od.cell(5 + i, 4, zone)

    if omit_sheet != "Integrated Drivers":
        dr = wb.create_sheet("Integrated Drivers")
        dr["B4"], dr["C4"] = "Driver ID", "Status"
        for i in range(n_drivers):
            dr.cell(5 + i, 2, f"DRV-{i:03d}")
            dr.cell(5 + i, 3, "ACTIVE")

    wb.save(path)
    return path
